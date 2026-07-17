import openpyxl
import re
import os
from datetime import datetime, date

path = r"C:\Users\simone.gizzi\Downloads\Per Atlantide\20260716_GM-MasterPlan_V16.xlsm"
wb = openpyxl.load_workbook(path, data_only=True, keep_vba=False)
ws = wb["Master"]

DATE_RE = re.compile(r"(\d{1,2})/(\d{1,2})/(\d{4})")


def to_date(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, str):
        m = DATE_RE.search(v.strip())
        if m:
            dd, mm, yyyy = int(m.group(1)), int(m.group(2)), int(m.group(3))
            try:
                return date(yyyy, mm, dd)
            except ValueError:
                return None
    return None


def sql_str(v):
    if v is None:
        return "NULL"
    s = str(v).replace("'", "''")
    return "'" + s + "'"


def sql_date(d):
    if d is None:
        return "NULL"
    return "'" + d.isoformat() + "'"


def sql_bool(b):
    return "true" if b else "false"


def sql_array(items):
    if not items:
        return "NULL"
    esc = [x.replace("'", "''").replace('"', '\\"') for x in items]
    inner = ",".join('"' + x + '"' for x in esc)
    return "'{" + inner + "}'"


CANONICI = {
    "Alesiani Veronica": ("Veronica", "Alesiani"),
    "Bontempi Lorenzo": ("Lorenzo", "Bontempi"),
    "Botezatu Daniel Andrei": ("Daniel Andrei", "Botezatu"),
    "Brunetti Eliseo": ("Eliseo", "Brunetti"),
    "Carriero Luciano": ("Luciano", "Carriero"),
    "Città Nello": ("Nello", "Città"),
    "Clementi Federica": ("Federica", "Clementi"),
    "Cozzi Andrea": ("Andrea", "Cozzi"),
    "Curzi Lorenzo": ("Lorenzo", "Curzi"),
    "Franceschi Marco": ("Marco", "Franceschi"),
    "Gramazio Gaetano": ("Gaetano", "Gramazio"),
    "Graziani Manuela": ("Manuela", "Graziani"),
    "Lo Iacono Gianluca": ("Gianluca", "Lo Iacono"),
    "Lopoio Enrico": ("Enrico", "Lopoio"),
    "Magnarelli Marco": ("Marco", "Magnarelli"),
    "Malatesta Giovanna": ("Giovanna", "Malatesta"),
    "Matera Claudia": ("Claudia", "Matera"),
    "Palestrini Cristian": ("Cristian", "Palestrini"),
    "Patti Stefano": ("Stefano", "Patti"),
    "Raffaelli Silvia": ("Silvia", "Raffaelli"),
    "Severoni Fabrizio": ("Fabrizio", "Severoni"),
    "Zottola Giada": ("Giada", "Zottola"),
    "Andrea Salzano": ("Andrea", "Salzano"),
    "Emanuele Bondatti": ("Emanuele", "Bondatti"),
}


def split_persona(raw_token):
    t = raw_token.strip()
    if not t:
        return None
    if t in CANONICI:
        return CANONICI[t]
    parts = t.split(" ")
    if len(parts) == 1:
        return (parts[0], "")
    return (parts[-1], " ".join(parts[:-1]))


PRIORITA_MAP = {"1": "Alta", "2": "Media", "3": "Bassa", "4": "Bassa"}

EFFORT_MAP = {
    "0-5": "0-5",
    "6-15": "11-20",
    "16-30": "21-60",
    "31-50": "21-60",
    "50+": "61-100",
}

progetti_rows = []
allocazioni_rows = []
persone_usate = {}
priorita_anomale = []
effort_map_log = {}

max_row = ws.max_row
seq = 0
for row in range(3, max_row + 1):
    area = ws.cell(row=row, column=2).value
    stato = ws.cell(row=row, column=7).value
    if area is None and stato is None:
        continue
    seq += 1
    id_progetto = "MIG-%03d" % seq

    prodotto_raw = ws.cell(row=row, column=3).value
    prodotto = [p.strip() for p in prodotto_raw.split(",")] if prodotto_raw else []
    macro_attivita = ws.cell(row=row, column=4).value
    descrizione = ws.cell(row=row, column=5).value
    priorita_raw = ws.cell(row=row, column=6).value
    priorita_key = str(priorita_raw).strip() if priorita_raw is not None else None
    note = ws.cell(row=row, column=55).value or ""
    if priorita_key in PRIORITA_MAP:
        priorita = PRIORITA_MAP[priorita_key]
    elif priorita_key:
        priorita = None
        note = (note + " [Priorità originale: " + priorita_key + " — da rivedere manualmente]").strip()
        priorita_anomale.append((id_progetto, priorita_key))
    else:
        priorita = None

    jira_raw = ws.cell(row=row, column=8).value
    jira_code = jira_raw.strip() if jira_raw and jira_raw.strip() not in ("-", "") else None
    fonte = "Jira" if jira_code else "Manuale"

    owner_progetto = ws.cell(row=row, column=34).value
    referenti = ws.cell(row=row, column=33).value
    action_owner = ws.cell(row=row, column=43).value
    last_action = ws.cell(row=row, column=41).value
    to_do = ws.cell(row=row, column=42).value

    deadline = to_date(ws.cell(row=row, column=44).value)
    deadline_r_raw = ws.cell(row=row, column=45).value
    if isinstance(deadline_r_raw, str) and "," in deadline_r_raw:
        ultima = deadline_r_raw.split(",")[-1]
        due_ripianificata = to_date(ultima)
    else:
        due_ripianificata = to_date(deadline_r_raw)
    deadline_chiusura = to_date(ws.cell(row=row, column=46).value)
    go_live = to_date(ws.cell(row=row, column=47).value)
    quarter = ws.cell(row=row, column=50).value

    critico_raw = ws.cell(row=row, column=54).value
    critico = bool(critico_raw and str(critico_raw).strip().upper() == "SI")

    effort_storico = ws.cell(row=row, column=35).value
    effort_storico_key = str(effort_storico).strip() if effort_storico else None
    effort_nuovo = EFFORT_MAP.get(effort_storico_key)
    if effort_storico_key:
        effort_map_log[effort_storico_key] = effort_map_log.get(effort_storico_key, 0) + 1

    progetti_rows.append(dict(
        id_progetto=id_progetto, area=area, prodotto=prodotto, macro_attivita=macro_attivita,
        descrizione=descrizione, priorita=priorita, stato=stato, jira_code=jira_code,
        fonte=fonte, owner_progetto=owner_progetto, referenti=referenti, action_owner=action_owner,
        last_action=last_action, to_do=to_do, due_date=deadline, due_date_ripianificata=due_ripianificata,
        deadline_chiusura=deadline_chiusura, go_live_date=go_live, quarter_pianificato=quarter,
        critico_per_business=critico, note=note or None, effort=effort_nuovo, effort_storico=effort_storico_key,
    ))

    risorse_raw = ws.cell(row=row, column=36).value
    if risorse_raw:
        tokens = [t.strip() for t in re.split(r"[,\n]", risorse_raw) if t.strip()]
        for tok in tokens:
            parsed = split_persona(tok)
            if parsed is None:
                continue
            nome, cognome = parsed
            persone_usate[(nome, cognome)] = True
            allocazioni_rows.append((id_progetto, nome, cognome))

out_path = r"C:\Users\simone.gizzi\OneDrive - DGS SpA\Desktop\Atlantide\scripts\import_masterplan.sql"
os.makedirs(os.path.dirname(out_path), exist_ok=True)

with open(out_path, "w", encoding="utf-8") as f:
    f.write("-- Import Anagrafica_Risorse\n")
    for (nome, cognome) in sorted(persone_usate.keys()):
        f.write(
            "insert into anagrafica_risorse (nome, cognome) values (" + sql_str(nome) + ", " + sql_str(cognome) +
            ") on conflict (nome, cognome) do nothing;\n"
        )

    f.write("\n-- Import Progetti\n")
    for p in progetti_rows:
        cols = (
            "id_progetto, area, prodotto, macro_attivita, descrizione, priorita, stato, "
            "jira_code, fonte, owner_progetto, referenti, action_owner, last_action, to_do, due_date, "
            "due_date_ripianificata, deadline_chiusura, go_live_date, quarter_pianificato, critico_per_business, "
            "note, effort, effort_storico"
        )
        vals = ", ".join([
            sql_str(p["id_progetto"]), sql_str(p["area"]), sql_array(p["prodotto"]), sql_str(p["macro_attivita"]),
            sql_str(p["descrizione"]), sql_str(p["priorita"]), sql_str(p["stato"]), sql_str(p["jira_code"]),
            sql_str(p["fonte"]), sql_str(p["owner_progetto"]), sql_str(p["referenti"]), sql_str(p["action_owner"]),
            sql_str(p["last_action"]), sql_str(p["to_do"]), sql_date(p["due_date"]), sql_date(p["due_date_ripianificata"]),
            sql_date(p["deadline_chiusura"]), sql_date(p["go_live_date"]), sql_str(p["quarter_pianificato"]),
            sql_bool(p["critico_per_business"]), sql_str(p["note"]), sql_str(p["effort"]), sql_str(p["effort_storico"]),
        ])
        f.write("insert into progetti (" + cols + ") values (" + vals + ") on conflict (id_progetto) do nothing;\n")

    f.write("\n-- Import Allocazioni\n")
    for id_progetto, nome, cognome in allocazioni_rows:
        f.write(
            "insert into allocazioni (id_progetto, nome, cognome) values (" +
            sql_str(id_progetto) + ", " + sql_str(nome) + ", " + sql_str(cognome) + ");\n"
        )

print("File scritto:", out_path)
print("Righe progetti:", len(progetti_rows))
print("Righe allocazioni:", len(allocazioni_rows))
print("Persone distinte:", len(persone_usate))
print()
print("Log mappatura Effort storico -> nuovo:")
for k, v in effort_map_log.items():
    print(" ", repr(k), "->", repr(EFFORT_MAP.get(k)), "(", v, "righe)")
print()
print("Priorità anomale segnalate (non convertite):")
for idp, val in priorita_anomale:
    print(" ", idp, ":", repr(val))
